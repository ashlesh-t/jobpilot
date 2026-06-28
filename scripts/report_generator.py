"""Report generator — styled XLSX (Layer B output formatter, NO LLM here).

Reads pre-scored jobs (written by the /job-search skill into /tmp/jobpilot_scored.json)
and renders a styled, dated workbook to ~/.claude/job-hunt-ai/reports/YYYY-MM-DD-<slot>.xlsx.

Scoring, salary research, skill extraction and JD summarisation are done by Claude in
the /job-search skill and passed in via the input JSON — this script only formats.

Row colouring by match score: >=75 green, 60-74 yellow, else white.
Sorted by effective_score (score x location_weight), capped at the top 20.

Usage:
  python3 scripts/report_generator.py \
      --input /tmp/jobpilot_scored.json \
      --output ~/.claude/job-hunt-ai/reports/2026-06-27-morning.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

IST = timezone(timedelta(hours=5, minutes=30))
TOP_N = 20

# (header, job-dict key, width). Order defines column order.
COLUMNS = [
    ("#", "_row", 5),
    ("Company", "company", 22),
    ("Role", "role", 30),
    ("Location", "location", 20),
    ("Exp Required", "exp_required", 13),
    ("Must-Have Skills", "must_have_skills", 30),
    ("Nice-to-Have", "nice_to_have", 24),
    ("Degree Required", "degree_required", 16),
    ("Match Score", "score", 11),
    ("Why", "why", 42),
    ("Matched Skills", "matched_skills", 28),
    ("Missing Skills", "missing_skills", 26),
    ("Market Salary", "market_salary", 15),
    ("Your Demand", "your_demand", 13),
    ("Salary Source", "salary_source", 14),
    ("Apply Link", "application_url", 40),
    ("Apply Type", "apply_type", 14),
    ("Source Board", "source_board", 14),
    ("Posted Date", "posted_date", 13),
    ("JD Summary", "jd_summary", 55),
    ("Job ID", "job_id", 18),
]

HEADER_FILL = "1F4E79"
GREEN_FILL = "C6EFCE"
YELLOW_FILL = "FFEB9C"
GREEN_FONT = "006100"
YELLOW_FONT = "9C6500"


def jobpilot_dir() -> Path:
    raw = os.environ.get("JOBPILOT_DIR", "~/.claude/job-hunt-ai")
    return Path(os.path.expanduser(raw))


def load_json(path, default):
    try:
        return json.loads(Path(path).read_text())
    except Exception:
        return default


def slot_now() -> str:
    hour = datetime.now(IST).hour
    if hour < 12:
        return "morning"
    if hour < 17:
        return "afternoon"
    return "evening"


def classify_apply_url(url: str) -> str:
    u = (url or "").lower()
    if not u:
        return "none"
    if "greenhouse.io" in u or "greenhouse" in u:
        return "greenhouse"
    if "lever.co" in u:
        return "lever"
    if "ashbyhq.com" in u:
        return "ashby"
    if "workday" in u or "myworkdayjobs" in u:
        return "workday"
    if "docs.google.com/forms" in u or "forms.gle" in u:
        return "⚠️ google_form"
    if "internshala.com" in u:
        return "internshala"
    if "naukri.com" in u:
        return "naukri"
    if "linkedin.com" in u:
        return "linkedin"
    if "cutshort.io" in u:
        return "cutshort"
    if "wellfound.com" in u or "angel.co" in u:
        return "wellfound"
    return "company"


def _join(value) -> str:
    """Render a list/str field to a clean cell string."""
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v).strip() for v in value if str(v).strip())
    return str(value).strip()


def _exp_required(job: dict) -> str:
    raw = (job.get("exp_required") or job.get("experience_req") or "").strip()
    if raw:
        return raw
    n = job.get("exp_req_years")
    if n is None or n == -1:
        return ""
    return "Fresher / 0-1 yr" if n == 0 else f"{n}+ yrs"


def _jd_summary(job: dict) -> str:
    val = job.get("jd_summary")
    if isinstance(val, (list, tuple)):
        return "\n".join(f"• {str(v).strip().lstrip('•- ')}" for v in val if str(v).strip())
    if val:
        return str(val).strip()
    # fallback: first ~2 sentences of jd_full
    jd = (job.get("jd_full") or "").replace("\n", " ")
    sentences = [s.strip() for s in jd.split(".") if len(s.strip()) > 25][:3]
    return "\n".join(f"• {s}" for s in sentences)


def effective_score(job: dict) -> float:
    score = float(job.get("score") or 0)
    return score * float(job.get("location_weight") or 1.0)


def cell_value(job: dict, key: str, row_index: int):
    if key == "_row":
        return row_index
    if key == "exp_required":
        return _exp_required(job)
    if key == "jd_summary":
        return _jd_summary(job)
    if key == "apply_type":
        return job.get("apply_type") or classify_apply_url(job.get("application_url", ""))
    if key == "score":
        try:
            return round(float(job.get("score") or 0))
        except (TypeError, ValueError):
            return 0
    if key in ("must_have_skills", "nice_to_have", "matched_skills", "missing_skills"):
        return _join(job.get(key) if key != "missing_skills"
                     else (job.get("missing_skills") or job.get("missing_keywords")))
    if key == "market_salary":
        return _join(job.get("market_salary") or job.get("salary_range"))
    if key == "your_demand":
        return _join(job.get("your_demand") or job.get("demand_estimate"))
    if key == "why":
        return _join(job.get("why") or job.get("why_score"))
    return _join(job.get(key))


def build_workbook(jobs: list):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    jobs = sorted(jobs, key=effective_score, reverse=True)[:TOP_N]

    wb = Workbook()
    ws = wb.active
    ws.title = "JobPilot"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top")

    # Header row
    for col_idx, (header, _key, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(60, width))
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    wrap_keys = {"why", "jd_summary", "must_have_skills", "nice_to_have",
                 "matched_skills", "missing_skills"}
    center_keys = {"_row", "score", "exp_required", "posted_date"}

    for r, job in enumerate(jobs, start=2):
        row_index = r - 1
        score = 0
        try:
            score = round(float(job.get("score") or 0))
        except (TypeError, ValueError):
            score = 0
        low_conf = (job.get("score_confidence") == "low")

        fill = None
        font_color = None
        if score >= 75:
            fill, font_color = GREEN_FILL, GREEN_FONT
        elif score >= 60:
            fill, font_color = YELLOW_FILL, YELLOW_FONT

        for col_idx, (_header, key, _width) in enumerate(COLUMNS, start=1):
            value = cell_value(job, key, row_index)
            c = ws.cell(row=r, column=col_idx, value=value)
            if key in wrap_keys:
                c.alignment = wrap
            elif key in center_keys:
                c.alignment = center
            else:
                c.alignment = Alignment(vertical="top")

            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
                if font_color:
                    c.font = Font(color=font_color)
            if key == "score":
                c.number_format = "0"
                c.font = Font(bold=True, color=font_color or "000000",
                              italic=low_conf)
            if key == "application_url" and value:
                c.hyperlink = value
                c.font = Font(color="0563C1", underline="single")
            if key == "apply_type" and isinstance(value, str) and value.startswith("⚠️"):
                c.font = Font(color="C00000", bold=True)

        ws.row_dimensions[r].height = 60

    return wb, len(jobs)


def write_csv_fallback(jobs: list, out_path: Path) -> Path:
    """Plain-CSV fallback used only if openpyxl is unavailable."""
    import csv
    out_path = out_path.with_suffix(".csv")
    jobs = sorted(jobs, key=effective_score, reverse=True)[:TOP_N]
    headers = [h for h, _k, _w in COLUMNS]
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        for i, job in enumerate(jobs, start=1):
            w.writerow([cell_value(job, key, i) for _h, key, _w in COLUMNS])
    return out_path


def write_report(jobs: list, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb, _n = build_workbook(jobs)
    except ImportError:
        print("[report] openpyxl not installed — writing CSV fallback "
              "(run: pip install openpyxl).", file=sys.stderr)
        return write_csv_fallback(jobs, out_path)
    wb.save(str(out_path))
    return out_path


def default_output() -> Path:
    fname = f"{datetime.now(IST).strftime('%Y-%m-%d')}-{slot_now()}.xlsx"
    return jobpilot_dir() / "reports" / fname


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="/tmp/jobpilot_scored.json",
                    help="Scored jobs JSON (falls back to /tmp/jobpilot_filtered.json).")
    ap.add_argument("--output", default="", help="Output .xlsx path.")
    args = ap.parse_args()

    jobs = load_json(args.input, None)
    if jobs is None:
        jobs = load_json("/tmp/jobpilot_filtered.json", [])
        print(f"[report] {args.input} not found — using /tmp/jobpilot_filtered.json",
              file=sys.stderr)

    missing_url_ids = [j.get("job_id", "?") for j in jobs if not j.get("application_url")]
    if missing_url_ids:
        print(f"[report] {len(missing_url_ids)} rows missing apply link: "
              f"{', '.join(missing_url_ids[:10])}"
              + (" ..." if len(missing_url_ids) > 10 else ""), file=sys.stderr)

    out = Path(os.path.expanduser(args.output)) if args.output else default_output()
    path = write_report(jobs, out)
    shown = min(len(jobs), TOP_N)
    print(f"Report written: {shown} jobs (of {len(jobs)}) -> {path}")


if __name__ == "__main__":
    main()
